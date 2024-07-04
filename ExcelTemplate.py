import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from openpyxl.drawing.image import Image
from openpyxl.worksheet.page import PageMargins
from datetime import datetime

# สร้าง DataFrame โดยไม่มีคอลัมน์ "Name 29"
data = {
    "Extruder Condition": ["Time"],
    "Name 2": ["Speed Screw Feeder(tons/hr)"],
    "Name 3": ["Feed Rate(kgs/hr)"],
    "Name 4": ["Ext.Load(%)"],
    "Name 5": ["Extruder Speed(rpm)"],
    "Name 6": ["Knife Speed(rpm)"],
    "Name 7": ["Downspout Temp(oC)"],
    "Name 8": ["Temp.Barrel1(oC)"],
    "Name 9": ["Temp.Barrel2(oC)"],
    "Name 10": ["Temp.Barrel3(oC)"],
    "Name 11": ["Pre-cond.Steam(kgs/hr)"],
    "Name 12": ["Pre-cond.Water(L/hr)"],
    "Name 13": ["Color Flow in Cond.(kgs/hr.)"],
    "Name 14": ["Fat Flow in Cond.(kgs/min.)"],
    "Name 15": ["Stream Header(bar)"],
    "Dryer Temp. (oC)": ["Dryer Temp Zone 1"],
    "Name 17": ["Dryer Temp Zone 2"],
    "Time(min.)": ["Deck No. 1-5"],
    "Fan Speed(Hz)": ["Recir. Blow 1"],
    "Name 20": ["Recir. Blow 2"],
    "": ["Capacity (kg/hr)"],
    "ข้อมูลอื่นๆ": ["ความชื้น Dyer"],
    "Name 23": ["ความชื้นCooler"],
    "Name 24": ["อุณหภูมิห้อง"],
    "Name 25": ["ความชื้นสัมพัทธ์"],
    "Name 26": ["ความชื้นหน้าเครื่อง"],
    "Name 27": ["กายภาพและปัญหาที่พบ"],
}

df = pd.DataFrame(data)

# กำหนดที่อยู่และชื่อไฟล์ที่จะบันทึก
file_path = os.path.expanduser("~D:/python/dash/my_plc_report_app/plc")  # เปลี่ยนเป็นที่อยู่ไดเร็กทอรีที่มีสิทธิ์เขียน
file_name = "Extruder1_Report.xlsx"
excel_file = os.path.join(file_path, file_name)

# บันทึก DataFrame เป็นไฟล์ Excel
df.to_excel(excel_file, index=False, sheet_name="Extruder1_Report")

# โหลดไฟล์ Excel เพื่อแก้ไข
wb = load_workbook(excel_file)
ws = wb["Extruder1_Report"]

# แทรกสามแถวที่ด้านบนสุด
ws.insert_rows(1, 3)

# แทรกคอลัมน์ใหม่ก่อนคอลัมน์ A
ws.insert_cols(1)

# รวมเซลล์ A1 ถึง A3
ws.merge_cells('A1:A3')

# เพิ่มรูปภาพในเซลล์ A1
img_path = "D:/python/dash/my_plc_report_app/plc/images/logo.png"  # เปลี่ยนเป็นที่อยู่ของรูปภาพ
img = Image(img_path)
img.width = 90  # กำหนดความกว้างของรูปภาพ
img.height = 80  # กำหนดความสูงของรูปภาพ
img.anchor = 'A1'
ws.add_image(img)

img_path = "D:/python/dash/my_plc_report_app/plc/images/ผู้ตรวจสอบ.png"  # เปลี่ยนเป็นที่อยู่ของรูปภาพ
img = Image(img_path)
img.width = 150  # กำหนดความกว้างของรูปภาพ
img.height = 80  # กำหนดความสูงของรูปภาพ
img.anchor = 'W22'
ws.add_image(img)

img_path = "D:/python/dash/my_plc_report_app/plc/images/ผู้บันทึก.png"  # เปลี่ยนเป็นที่อยู่ของรูปภาพ
img = Image(img_path)
img.width = 150  # กำหนดความกว้างของรูปภาพ
img.height = 80  # กำหนดความสูงของรูปภาพ
img.anchor = 'B22'
ws.add_image(img)

# ปรับตำแหน่ง anchor ของรูปภาพให้อยู่ตรงกลางช่อง A1
ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
# เพิ่มข้อความในเซลล์ B1 และจัดให้กลางช่อง
ws['B1'] = 'บริษัท บลูฟาโล่ เพ็ทแคร์ จำกัด'
ws['B1'].alignment = Alignment(horizontal="center", vertical="center")
ws['B1'].font = Font(name='TH SarabunPSK', size=20, bold=True)  # ตั้งขนาดตัวอักษรเป็น 20, ฟอนต์เป็น TH SarabunPSK และตัวหนา และสีฟอนต์เป็นสีแดง

ws['B2'] = 'บันทึกการทำงานของ Extruder และ Dryer'
ws['B2'].alignment = Alignment(horizontal="center", vertical="center")
ws['B2'].font = Font(name='TH SarabunPSK', size=20, bold=True)  # ตั้งขนาดตัวอักษรเป็น 20, ฟอนต์เป็น TH SarabunPSK และตัวหนา และสีฟอนต์เป็นสีแดง


ws.merge_cells('B3:D3')
ws['B3'] = 'User: '
ws['B3'].alignment = Alignment(horizontal="center", vertical="center")
ws['B3'].font = Font(name='TH SarabunPSK', size=12, bold=True, color="FF0000")


ws.merge_cells('E3:G3')
ws['E3'] = 'Dynamic Name'
ws['E3'].alignment = Alignment(horizontal="center", vertical="center")
ws['E3'].font = Font(name='TH SarabunPSK', size=12, bold=True, color="FF0120")


ws.merge_cells('K3:M3')
ws['K3'] = 'Date: '
ws['K3'].alignment = Alignment(horizontal="center", vertical="center")
ws['K3'].font = Font(name='TH SarabunPSK', size=12, bold=True, color="FF0000")

ws.merge_cells('U3:W3')
ws['U3'] = 'Shift: '
ws['U3'].alignment = Alignment(horizontal="center", vertical="center")
ws['U3'].font = Font(name='TH SarabunPSK', size=12, bold=True, color="FF0000")

ws.merge_cells('N3:Q3')
ws['N3'] = datetime.today().strftime('%Y-%m-%d')  # วันที่จริงเมื่อสร้างรายงาน
ws['N3'].alignment = Alignment(horizontal="center", vertical="center")
ws['N3'].font = Font(name='TH SarabunPSK', size=12, bold=True, color="FF0000")

ws.merge_cells('X3:AB3')
ws['X3'] = '8:00-16:30'  # ปรับเปลี่ยนช่วงเวลาตามที่ต้องการ
ws['X3'].alignment = Alignment(horizontal="center", vertical="center")
ws['X3'].font = Font(name='TH SarabunPSK', size=12, bold=True, color="FF0120")


ws['A6'] = 'Timestamp:'
ws['A6'].alignment = Alignment(horizontal="left", vertical="center")
ws['A6'].font = Font(name='TH SarabunPSK', size=10, bold=True)

ws['A7'] = 'Extruder No:'
ws['A7'].alignment = Alignment(horizontal="left", vertical="center")
ws['A7'].font = Font(name='TH SarabunPSK', size=10, bold=True)

ws['A8'] = 'Name Food:'
ws['A8'].alignment = Alignment(horizontal="left", vertical="center")
ws['A8'].font = Font(name='TH SarabunPSK', size=10, bold=True)

ws['A9'] = 'Code Lot:'
ws['A9'].alignment = Alignment(horizontal="left", vertical="center")
ws['A9'].font = Font(name='TH SarabunPSK', size=10, bold=True)

ws['A10'] = 'Shape:'
ws['A10'].alignment = Alignment(horizontal="left", vertical="center")
ws['A10'].font = Font(name='TH SarabunPSK', size=10, bold=True)

ws['A11'] = 'Number Dir:'
ws['A11'].alignment = Alignment(horizontal="left", vertical="center")
ws['A11'].font = Font(name='TH SarabunPSK', size=10, bold=True)

ws['A12'] = 'Size:'
ws['A12'].alignment = Alignment(horizontal="left", vertical="center")
ws['A12'].font = Font(name='TH SarabunPSK', size=10, bold=True)

ws['A13'] = 'Number of Blades:'
ws['A13'].alignment = Alignment(horizontal="left", vertical="center")
ws['A13'].font = Font(name='TH SarabunPSK', size=10, bold=True)

ws['A14'] = 'Color:'
ws['A14'].alignment = Alignment(horizontal="left", vertical="center")
ws['A14'].font = Font(name='TH SarabunPSK', size=10, bold=True)

ws['A15'] = 'Oil:'
ws['A15'].alignment = Alignment(horizontal="left", vertical="center")
ws['A15'].font = Font(name='TH SarabunPSK', size=10, bold=True)

ws['A16'] = 'Lot color oil:'
ws['A16'].alignment = Alignment(horizontal="left", vertical="center")
ws['A16'].font = Font(name='TH SarabunPSK', size=10, bold=True)

ws['A17'] = 'Weight color oil:'
ws['A17'].alignment = Alignment(horizontal="left", vertical="center")
ws['A17'].font = Font(name='TH SarabunPSK', size=10, bold=True)

# รวมเซลล์ B1 ถึง X1 และ B2 ถึง X2
ws.merge_cells('B1:AB1')
ws.merge_cells('B2:AB2')
ws.merge_cells('A4:A5')

# เพิ่มการรวมเซลล์สำหรับหัวตาราง
ws.merge_cells('B4:P4')
ws.merge_cells('Q4:R4')
ws.merge_cells('T4:U4')
ws.merge_cells('W4:AB4')
ws.merge_cells('AB6:AB19')

# ตั้งค่าฟอนต์และการจัดแนวสำหรับหัวตาราง
header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # สีเหลือง

for col in range(2, len(data) + 1):
    cell = ws.cell(row=4, column=col)
    cell.font = Font(name='TH SarabunPSK', size=10, bold=True, color="FF0000")
    cell.alignment = Alignment(horizontal="center", vertical="center")
 

# จัดแนวข้อความในเซลล์ในช่วงที่กำหนดและหมุนข้อความในแนวตั้ง
for row in ws.iter_rows(min_row=5, max_row=len(df) + 4, min_col=2, max_col=len(df.columns) + 1):
    for cell in row:
        cell.alignment = Alignment(horizontal="center", vertical="center", textRotation=90)
        cell.font = Font(name='TH SarabunPSK', size=10, bold=True)  # ตั้งฟอนต์เป็น TH SarabunPSK และตัวหนา
        thin_border = Border(left=Side(style='thin', color="000000"),
                             right=Side(style='thin', color="000000"),
                             top=Side(style='thin', color="000000"),
                             bottom=Side(style='thin', color="000000"))

# ปรับขนาดคอลัมน์ A ตั้งแต่แถวที่ 5 ให้มีขนาดประมาณ 50 หน่วยความกว้าง (ประมาณ 350 พิกเซล)
ws.column_dimensions['A'].width = 20

# ปรับขนาดคอลัมน์ B ถึง P ตั้งแต่แถวที่ 5 (รวมคอลัมน์ใหม่)
for col in range(3, 17):  # คอลัมน์ C ถึง P คือคอลัมน์ที่ 2 ถึง 16
    col_letter = chr(64 + col)  # แปลงตัวเลขเป็นตัวอักษร เช่น 2 -> B, 3 -> C, ...
    ws.column_dimensions[col_letter].width = 3.9

# ปรับขนาดคอลัมน์ Q และ R ให้มีความกว้าง 5.2
ws.column_dimensions['B'].width = 6
ws.column_dimensions['Q'].width = 6
ws.column_dimensions['R'].width = 6
ws.column_dimensions['T'].width = 5
ws.column_dimensions['U'].width = 5
ws.column_dimensions['V'].width = 5
ws.column_dimensions['W'].width = 5
ws.column_dimensions['X'].width = 5
ws.column_dimensions['Y'].width = 5
ws.column_dimensions['Z'].width = 5
ws.column_dimensions['AA'].width = 5
ws.column_dimensions['AB'].width = 5
ws.column_dimensions['AC'].width = 5  # คอลัมน์ใหม่
ws.column_dimensions['S'].width = 6
ws.column_dimensions['C'].width = 5

# ปรับกระดาษเป็นแนวนอนและขนาด A4
ws.page_setup.orientation = 'landscape'
ws.page_setup.paperSize = ws.PAPERSIZE_A4

# ตั้งค่าหน้ากระดาษให้พอดีกับหน้ากระดาษ A4
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 1

# ตั้งค่าหน้ากระดาษให้มีระยะขอบตามที่กำหนด
ws.page_margins = PageMargins(top=1.91/2.54, bottom=1.91/2.54, left=0.64/2.54, right=0.64/2.54, header=0.76/2.54, footer=0.76/2.54)

# เพิ่มข้อความในเซลล์ A4 และจัดให้กลางช่อง
ws['A4'] = 'ชื่ออาหาร / ข้อมูล Die'
ws['A4'].alignment = Alignment(horizontal="center", vertical="center")
ws['A4'].font = Font(name='TH SarabunPSK', size=13, bold=True)  # ตั้งฟอนต์เป็น TH SarabunPSK ขนาด 10 และตัวหนา

# กำหนดเส้นขอบให้ทุกเซลล์ตั้งแต่ A4 ถึง AC5 และให้เส้นขอบเป็นสีดำ
thick_border = Border(left=Side(style='thin', color="000000"),
                      right=Side(style='thin', color="000000"),
                      top=Side(style='thin', color="000000"),
                      bottom=Side(style='thin', color="000000"))

for row in ws.iter_rows(min_row=4, max_row=5, min_col=1, max_col=len(data) + 1):
    for cell in row:
        cell.border = thick_border
        cell.font = Font(name='TH SarabunPSK', size=10, bold=True)  # ตั้งฟอนต์เป็น TH SarabunPSK ขนาด 10 และตัวหนา

# กำหนดเส้นขอบบางและสีดำสำหรับทุกเซลล์ตั้งแต่ A6 ถึง AC20
thin_border = Border(
    left=Side(style='thin', color="000000"),
    right=Side(style='thin', color="000000"),
    top=Side(style='thin', color="000000"),
    bottom=Side(style='thin', color="000000")
)
for row in ws.iter_rows(min_row=6, max_row=19, min_col=1, max_col=28):  # A=1, B=2, ..., AC=29
    for cell in row:
        cell.border = thin_border
        cell.font = Font(name='TH SarabunPSK', size=10, bold=True)  # ตั้งฟอนต์เป็น TH SarabunPSK ขนาด 10 และตัวหนา

# บันทึกการเปลี่ยนแปลงในไฟล์ Excel
wb.save(excel_file)
